"""Tests for filter and role parameters in report download API."""
from django.test import TestCase
from django.contrib.auth import get_user_model
from django.urls import reverse
from rest_framework.test import APIClient
from rest_framework import status
from core.models import LoginActivity
from datetime import timedelta
from django.utils import timezone
import io
from openpyxl import load_workbook

User = get_user_model()


class ReportDownloadFilterParameterTests(TestCase):
    """Test the 'filter' parameter for report download."""

    def setUp(self):
        self.admin_user = User.objects.create_superuser(
            username='admin',
            email='admin@example.com',
            password='adminpass123'
        )
        self.regular_user1 = User.objects.create_user(
            username='regular1',
            email='regular1@example.com',
            password='testpass123'
        )
        self.regular_user2 = User.objects.create_user(
            username='regular2',
            email='regular2@example.com',
            password='testpass123'
        )
        self.client = APIClient()
        self.url = reverse('user:report-download')

        # Create login activities for regular users
        for i in range(3):
            LoginActivity.objects.create(
                user=self.regular_user1,
                ip_address=f'192.168.1.{i+1}',
                user_agent='Chrome',
                success=True,
                timestamp=timezone.now() - timedelta(days=i)
            )
        for i in range(2):
            LoginActivity.objects.create(
                user=self.regular_user2,
                ip_address=f'192.168.2.{i+1}',
                user_agent='Firefox',
                success=True,
                timestamp=timezone.now() - timedelta(days=i)
            )

    def test_filter_admin_only_returns_admin_users_data(self):
        """filter=admin_only should return data for admin users only."""
        self.client.force_authenticate(user=self.admin_user)
        response = self.client.get(
            self.url,
            {'mode': 'grouped', 'filter': 'admin_only'}
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        cells = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    cells.append(str(cell))
        cell_text = ' '.join(cells)

        # Should only include admin users - admin has no login data created
        # but the report should be generated for admin users
        self.assertIn('admin', cell_text)
        self.assertNotIn('regular1', cell_text)
        self.assertNotIn('regular2', cell_text)

    def test_filter_regular_users_returns_regular_users_data(self):
        """filter=regular_users should return data for regular users only."""
        self.client.force_authenticate(user=self.admin_user)
        response = self.client.get(
            self.url,
            {'mode': 'grouped', 'filter': 'regular_users'}
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        cells = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    cells.append(str(cell))
        cell_text = ' '.join(cells)

        # Should include regular users data
        self.assertIn('regular1', cell_text)
        self.assertIn('regular2', cell_text)
        # Header shows admin info but data is from regular users only
        self.assertIn('Regular Users', cell_text)

    def test_filter_me_returns_own_data(self):
        """filter=me should return only the requesting user's data."""
        self.client.force_authenticate(user=self.regular_user1)
        response = self.client.get(
            self.url,
            {'mode': 'individual', 'filter': 'me'}
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        cells = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    cells.append(str(cell))
        cell_text = ' '.join(cells)

        self.assertIn('regular1', cell_text)

    def test_filter_me_admin_returns_admin_own_data(self):
        """Admin using filter=me should get only their own data."""
        self.client.force_authenticate(user=self.admin_user)
        response = self.client.get(
            self.url,
            {'mode': 'individual', 'filter': 'me'}
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_invalid_filter_returns_error(self):
        """Invalid filter value should return 400."""
        self.client.force_authenticate(user=self.admin_user)
        response = self.client.get(
            self.url,
            {'mode': 'grouped', 'filter': 'invalid_filter'}
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('error', response.data)

    def test_filter_with_date_range(self):
        """filter should work properly with date range."""
        self.client.force_authenticate(user=self.admin_user)
        start_date = (
            timezone.now() - timedelta(days=5)
        ).strftime('%Y-%m-%d')
        end_date = timezone.now().strftime('%Y-%m-%d')
        response = self.client.get(
            self.url,
            {
                'mode': 'grouped',
                'filter': 'regular_users',
                'start_date': start_date,
                'end_date': end_date
            }
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)


class ReportDownloadRoleParameterTests(TestCase):
    """Test the 'role' parameter for report download."""

    def setUp(self):
        self.admin_user = User.objects.create_superuser(
            username='admin',
            email='admin@example.com',
            password='adminpass123'
        )
        self.regular_user = User.objects.create_user(
            username='regular',
            email='regular@example.com',
            password='testpass123'
        )
        self.client = APIClient()
        self.url = reverse('user:report-download')

        LoginActivity.objects.create(
            user=self.admin_user,
            ip_address='192.168.0.1',
            user_agent='Admin Browser',
            success=True,
            timestamp=timezone.now()
        )
        LoginActivity.objects.create(
            user=self.regular_user,
            ip_address='192.168.1.1',
            user_agent='Chrome',
            success=True,
            timestamp=timezone.now()
        )

    def test_role_admin_returns_only_admin_users(self):
        """role=admin should return data for admin users only."""
        self.client.force_authenticate(user=self.admin_user)
        response = self.client.get(
            self.url,
            {'mode': 'grouped', 'role': 'admin'}
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        cells = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    cells.append(str(cell))
        cell_text = ' '.join(cells)

        self.assertIn('admin', cell_text)
        self.assertNotIn('regular', cell_text)

    def test_role_regular_returns_only_regular_users(self):
        """role=regular should return data for regular users only."""
        self.client.force_authenticate(user=self.admin_user)
        response = self.client.get(
            self.url,
            {'mode': 'grouped', 'role': 'regular'}
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        cells = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    cells.append(str(cell))
        cell_text = ' '.join(cells)

        self.assertIn('regular', cell_text)
        self.assertIn('Role: Regular', cell_text)

    def test_invalid_role_returns_error(self):
        """Invalid role value should return 400."""
        self.client.force_authenticate(user=self.admin_user)
        response = self.client.get(
            self.url,
            {'mode': 'grouped', 'role': 'superuser'}
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('error', response.data)

    def test_role_requires_admin_permissions(self):
        """Regular user should not be able to use role parameter."""
        self.client.force_authenticate(user=self.regular_user)
        response = self.client.get(
            self.url,
            {'mode': 'grouped', 'role': 'admin'}
        )
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_role_with_individual_mode(self):
        """role parameter should work with individual mode for admin."""
        self.client.force_authenticate(user=self.admin_user)
        response = self.client.get(
            self.url,
            {'mode': 'individual', 'role': 'admin'}
        )
        # Should still work - role resolves users, individual mode takes first
        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_role_with_date_range(self):
        """role should work properly with date range."""
        self.client.force_authenticate(user=self.admin_user)
        start_date = (
            timezone.now() - timedelta(days=5)
        ).strftime('%Y-%m-%d')
        end_date = timezone.now().strftime('%Y-%m-%d')
        response = self.client.get(
            self.url,
            {
                'mode': 'grouped',
                'role': 'regular',
                'start_date': start_date,
                'end_date': end_date
            }
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)


class ReportDownloadFilterInExcelContentTests(TestCase):
    """Test that filter context appears correctly in Excel content."""

    def setUp(self):
        self.admin_user = User.objects.create_superuser(
            username='admin_user',
            email='admin@example.com',
            password='adminpass123'
        )
        self.regular_user = User.objects.create_user(
            username='test_regular',
            email='regular@example.com',
            password='testpass123'
        )
        self.client = APIClient()
        self.url = reverse('user:report-download')

        for i in range(3):
            LoginActivity.objects.create(
                user=self.regular_user,
                ip_address=f'192.168.1.{i+1}',
                user_agent='Firefox',
                success=True,
                timestamp=timezone.now() - timedelta(days=i)
            )

    def _get_excel_text(self, params):
        """Helper: get all text from an Excel response."""
        self.client.force_authenticate(user=self.admin_user)
        response = self.client.get(self.url, params)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        cells = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    cells.append(str(cell))
        return ' '.join(cells)

    # ── Filter context in all sections ──

    def test_excel_shows_filter_admin_only(self):
        """Excel should show 'Filter: Admin Only'."""
        text = self._get_excel_text(
            {'mode': 'grouped', 'filter': 'admin_only'}
        )
        self.assertIn('Filter: Admin Only', text)

    def test_excel_shows_filter_regular_users(self):
        """Excel should show 'Filter: Regular Users'."""
        text = self._get_excel_text(
            {'mode': 'grouped', 'filter': 'regular_users'}
        )
        self.assertIn('Filter: Regular Users', text)
        self.assertIn('Logged User:', text)

    def test_excel_shows_filter_all_users(self):
        """Excel should show 'Filter: All Users'."""
        text = self._get_excel_text(
            {'mode': 'grouped', 'filter': 'all'}
        )
        self.assertIn('Filter: All Users', text)

    def test_excel_shows_filter_me(self):
        """Excel should show 'Filter: My Data'."""
        text = self._get_excel_text(
            {'mode': 'individual', 'filter': 'me'}
        )
        self.assertIn('Filter: My Data', text)

    def test_excel_shows_role_admin(self):
        """Excel should show role in filter label."""
        text = self._get_excel_text(
            {'mode': 'grouped', 'role': 'admin'}
        )
        self.assertIn('Role: Admin', text)

    def test_excel_shows_role_regular(self):
        """Excel should show role in filter label."""
        text = self._get_excel_text(
            {'mode': 'grouped', 'role': 'regular'}
        )
        self.assertIn('Role: Regular', text)

    # ── Chart context format ──

    def test_excel_chart_context_individual_shows_filter_and_user(self):
        """Chart context in individual mode: 'Filter: X → Selected User: Y'."""
        text = self._get_excel_text(
            {'mode': 'individual', 'filter': 'regular_users'}
        )
        self.assertIn('Filter: Regular Users', text)
        self.assertIn('Selected User: test_regular', text)
        self.assertIn('Logged User:', text)

    def test_excel_chart_context_grouped_shows_filter_and_users(self):
        """Chart context in grouped mode: 'Filter: X → Users: ...'."""
        text = self._get_excel_text(
            {'mode': 'grouped', 'filter': 'admin_only'}
        )
        self.assertIn('Filter: Admin Only', text)
        self.assertIn('admin_user', text)

    def test_excel_chart_context_filter_all_individual(self):
        """Individual mode with filter=all shows 'Filter: All Users'."""
        text = self._get_excel_text(
            {'mode': 'individual', 'filter': 'all'}
        )
        self.assertIn('Filter: All Users', text)
        self.assertIn('Selected User: admin_user', text)
        # No 'Logged User' label since admin is viewing their own data
        # (admin_user is first user in 'all' filter result)

    def test_excel_chart_context_with_role(self):
        """Chart context should include role in label."""
        text = self._get_excel_text(
            {'mode': 'grouped', 'role': 'regular'}
        )
        self.assertIn('Filter:', text)
        self.assertIn('Role: Regular', text)
        self.assertIn('Logged User:', text)

    # ── Activities section filter context ──

    def test_excel_activities_shows_filter_context(self):
        """Activities section should show filter context."""
        text = self._get_excel_text(
            {'mode': 'individual', 'filter': 'regular_users'}
        )
        self.assertIn('Filter: Regular Users', text)
        self.assertIn('Selected User: test_regular', text)
        self.assertIn('Logged User:', text)

    # ── Summary section filter context ──

    def test_excel_summary_shows_filter_context(self):
        """Summary section should show filter context."""
        text = self._get_excel_text(
            {'mode': 'individual', 'filter': 'me'}
        )
        self.assertIn('Filter: My Data', text)
        self.assertIn('My Data', text)
