"""Tests for date filter offset bug - end_date should not shift to next day."""
from datetime import timedelta

from django.test import TestCase
from django.contrib.auth import get_user_model
from django.urls import reverse
from django.utils import timezone
from rest_framework.test import APIClient
from rest_framework import status
from core.models import LoginActivity
import io
from openpyxl import load_workbook

User = get_user_model()


class DateFilterEndDateOffsetTests(TestCase):
    """
    Tests to verify that end_date in date filters does NOT get shifted
    to the next day in the displayed period, and that data queries
    correctly include all records up to the end of the selected day.

    Bug: end_date + timedelta(days=1) caused:
      1. Display showing "2026-06-06" instead of "2026-06-05"
      2. DB query end bound being the next day at midnight instead of
         end of the selected day
    """

    def setUp(self):
        self.user = User.objects.create_user(
            username='testuser',
            email='test@example.com',
            password='testpass123'
        )

        self.admin_user = User.objects.create_superuser(
            username='admin',
            email='admin@example.com',
            password='adminpass123'
        )
        self.client = APIClient()
        self.dashboard_stats_url = reverse('user:dashboard-stats')
        self.report_url = reverse('user:report-download')

        # Create login activities at specific dates for precise testing
        self._create_test_data()

    def _create_test_data(self):
        """Create activities on specific dates for testing."""
        now = timezone.now()

        # Activity on the target end_date (e.g. "today")
        LoginActivity.objects.create(
            user=self.user,
            ip_address='192.168.1.1',
            user_agent='Browser Today',
            success=True,
            timestamp=now
        )

        # Activity yesterday
        LoginActivity.objects.create(
            user=self.user,
            ip_address='192.168.1.2',
            user_agent='Browser Yesterday',
            success=True,
            timestamp=now - timedelta(days=1)
        )

        # Activity 3 days ago
        LoginActivity.objects.create(
            user=self.user,
            ip_address='192.168.1.3',
            user_agent='Browser 3 Days Ago',
            success=True,
            timestamp=now - timedelta(days=3)
        )

        # Activity 10 days ago (well outside typical date range)
        LoginActivity.objects.create(
            user=self.user,
            ip_address='192.168.1.4',
            user_agent='Browser 10 Days Ago',
            success=True,
            timestamp=now - timedelta(days=10)
        )

    def test_report_excel_shows_correct_end_date_in_period(self):
        """
        Test that the Excel report's "Period:" line shows the correct
        end_date as selected, NOT end_date+1.

        User selects: 2026-05-29 to 2026-06-05
        Report should show: Period: 2026-05-29 to 2026-06-05
        NOT: Period: 2026-05-29 to 2026-06-06
        """
        self.client.force_authenticate(user=self.user)

        # Use specific dates to verify the display
        start_date = (timezone.now() - timedelta(days=10)).strftime('%Y-%m-%d')
        end_date = timezone.now().strftime('%Y-%m-%d')

        response = self.client.get(
            self.report_url,
            {
                'mode': 'individual',
                'start_date': start_date,
                'end_date': end_date
            }
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        # Parse the Excel content
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active

        # Read all cells to find the Period line
        cell_text = ''
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    cell_text += str(cell) + ' '

        # The period should show the user-selected end_date, not end_date+1
        expected_period = f"Period: {start_date} to {end_date}"
        self.assertIn(
            expected_period, cell_text,
            f"Expected period '{expected_period}' but got something else. "
            f"Full cell text: {cell_text}"
        )

    def test_dashboard_stats_end_date_includes_full_day_data(self):
        """
        Test that the dashboard stats correctly includes data from
        the full end_date day.

        When user selects end_date = today, activities from today
        should be included in the count.
        """
        self.client.force_authenticate(user=self.user)

        now = timezone.now()
        start_date = (now - timedelta(days=2)).strftime('%Y-%m-%d')
        end_date = now.strftime('%Y-%m-%d')

        # Expected: today's activity + yesterday's = 2 activities
        response = self.client.get(
            self.dashboard_stats_url,
            {
                'start_date': start_date,
                'end_date': end_date
            }
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response.data['total_logins'], 2,
            f"Expected 2 logins (today + yesterday), "
            f"got {response.data['total_logins']}"
        )
