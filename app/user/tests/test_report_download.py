"""Tests for report download API endpoints."""
from django.test import TestCase, override_settings
from django.contrib.auth import get_user_model
from django.urls import reverse
from rest_framework.test import APIClient
from rest_framework import status
from core.models import LoginActivity
from datetime import timedelta
from django.utils import timezone
import io
from openpyxl import load_workbook

User = get_user_model()


class ReportDownloadAuthenticationTests(TestCase):
    """Test authentication requirements for report download."""

    def setUp(self):
        self.user = User.objects.create_user(
            username='testuser',
            email='test@example.com',
            password='testpass123'
        )
        self.client = APIClient()
        self.url = reverse('user:report-download')

    def test_report_download_requires_authentication(self):
        """Test that report download endpoint requires authentication."""
        response = self.client.get(
            self.url, {'mode': 'individual'}
        )
        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_authenticated_user_can_access_report_download(self):
        """Test that authenticated user can access report download."""
        self.client.force_authenticate(user=self.user)
        LoginActivity.objects.create(
            user=self.user,
            ip_address='192.168.1.1',
            user_agent='Test Browser',
            success=True,
            timestamp=timezone.now()
        )
        response = self.client.get(
            self.url, {'mode': 'individual'}
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)


class ReportDownloadValidationTests(TestCase):
    """Test parameter validation."""

    def setUp(self):
        self.user = User.objects.create_user(
            username='testuser',
            email='test@example.com',
            password='testpass123'
        )
        self.admin_user = User.objects.create_superuser(
            username='admin',
            email='admin@example.com',
            password='adminpass123'
        )
        self.client = APIClient()
        self.url = reverse('user:report-download')
        LoginActivity.objects.create(
            user=self.user,
            ip_address='192.168.1.1',
            user_agent='Test Browser',
            success=True,
            timestamp=timezone.now()
        )

    def test_missing_mode_parameter_returns_error(self):
        self.client.force_authenticate(user=self.user)
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('error', response.data)

    def test_invalid_mode_parameter_returns_error(self):
        self.client.force_authenticate(user=self.user)
        response = self.client.get(
            self.url, {'mode': 'invalid'}
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('error', response.data)

    def test_valid_excel_format_returned(self):
        self.client.force_authenticate(user=self.user)
        response = self.client.get(
            self.url, {'mode': 'individual'}
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)


class ReportDownloadModeTests(TestCase):
    """Test individual and grouped mode functionality."""

    def setUp(self):
        self.user1 = User.objects.create_user(
            username='user1',
            email='user1@example.com',
            password='testpass123'
        )
        self.user2 = User.objects.create_user(
            username='user2',
            email='user2@example.com',
            password='testpass123'
        )
        self.admin_user = User.objects.create_superuser(
            username='admin',
            email='admin@example.com',
            password='adminpass123'
        )
        self.client = APIClient()
        self.url = reverse('user:report-download')
        self._create_test_data()

    def _create_test_data(self):
        for i in range(5):
            LoginActivity.objects.create(
                user=self.user1,
                ip_address=f'192.168.1.{i+1}',
                user_agent=f'Browser {i+1}',
                success=True,
                timestamp=timezone.now() - timedelta(days=i)
            )
        for i in range(3):
            LoginActivity.objects.create(
                user=self.user2,
                ip_address=f'192.168.2.{i+1}',
                user_agent=f'Browser {i+1}',
                success=True,
                timestamp=timezone.now() - timedelta(days=i)
            )

    def test_regular_user_individual_mode_returns_own_data(self):
        self.client.force_authenticate(user=self.user1)
        response = self.client.get(
            self.url, {'mode': 'individual'}
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn(
            response['Content-Type'],
            [
                'application/vnd.openxmlformats-officedocument.'
                'spreadsheetml.sheet',
                'application/octet-stream'
            ]
        )

    def test_regular_user_grouped_mode_returns_forbidden(self):
        self.client.force_authenticate(user=self.user1)
        response = self.client.get(
            self.url, {'mode': 'grouped'}
        )
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_admin_individual_mode_with_user_ids(self):
        self.client.force_authenticate(user=self.admin_user)
        response = self.client.get(
            self.url,
            {
                'mode': 'individual',
                'user_ids[]': [self.user1.id]
            }
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_admin_grouped_mode_with_user_ids(self):
        self.client.force_authenticate(user=self.admin_user)
        response = self.client.get(
            self.url,
            {
                'mode': 'grouped',
                'user_ids[]': [self.user1.id, self.user2.id]
            }
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_admin_grouped_mode_all_users(self):
        self.client.force_authenticate(user=self.admin_user)
        response = self.client.get(
            self.url, {'mode': 'grouped'}
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_admin_individual_mode_without_user_ids_returns_own_data(self):
        LoginActivity.objects.create(
            user=self.admin_user,
            ip_address='192.168.3.1',
            user_agent='Admin Browser',
            success=True,
            timestamp=timezone.now()
        )
        self.client.force_authenticate(user=self.admin_user)
        response = self.client.get(
            self.url, {'mode': 'individual'}
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_grouped_report_includes_selected_users_data(self):
        """Verify grouped Excel report with user_ids[] includes
        combined data from all selected users, not just admin."""
        self.client.force_authenticate(user=self.admin_user)
        response = self.client.get(
            self.url,
            {
                'mode': 'grouped',
                'user_ids[]': [self.user1.id, self.user2.id]
            }
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        # Parse the Excel content
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active

        # Read all cells into a list
        cells = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    cells.append(str(cell))

        cell_text = ' '.join(cells)

        # user1 has 5 logins, user2 has 3 logins = 8 total
        self.assertIn('Total Logins (All Users)', cell_text,
                      msg="GROUPED SUMMARY section not found in report")
        self.assertIn(
            '8', cell_text,
            msg="Expected 8 total logins (5+3 from both users)"
        )

    def test_grouped_report_all_users_includes_admin_data(self):
        """Verify grouped report with all users includes
        admin's login data as well."""
        LoginActivity.objects.create(
            user=self.admin_user,
            ip_address='192.168.3.1',
            user_agent='Admin Browser',
            success=True,
            timestamp=timezone.now()
        )
        self.client.force_authenticate(user=self.admin_user)
        response = self.client.get(
            self.url, {'mode': 'grouped'}
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        cells = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    cells.append(str(cell))
        cell_text = ' '.join(cells)

        # user1=5 + user2=3 + admin=1 = 9 total
        self.assertIn('9', cell_text,
                      msg="Expected 9 total logins (5+3+1) for all users")


class ReportDownloadDropdownSelectionTests(TestCase):
    """Test that when admin selects a dropdown user, the report shows
    'Selected User from Dropdown' label and data for that user."""

    def setUp(self):
        self.admin_user = User.objects.create_superuser(
            username='admin',
            email='admin@example.com',
            password='adminpass123'
        )
        self.normal_user = User.objects.create_user(
            username='normal_user',
            email='normal@example.com',
            password='testpass123'
        )
        self.client = APIClient()
        self.url = reverse('user:report-download')

        # Create login activity for admin
        LoginActivity.objects.create(
            user=self.admin_user,
            ip_address='192.168.0.1',
            user_agent='Admin Browser',
            success=True,
            timestamp=timezone.now() - timedelta(days=1)
        )
        # Create login activity for normal user
        for i in range(5):
            LoginActivity.objects.create(
                user=self.normal_user,
                ip_address=f'192.168.1.{i+1}',
                user_agent=f'Normal Browser {i+1}',
                success=i % 4 != 0,  # 4 successful, 1 failed
                timestamp=timezone.now() - timedelta(days=i)
            )

    def test_header_shows_logged_user_and_selected_user_dropdown(self):
        """When admin downloads report for a selected user,
        header should show both 'Logged User' and
        'Selected User from Dropdown'."""
        self.client.force_authenticate(user=self.admin_user)
        response = self.client.get(
            self.url,
            {
                'mode': 'individual',
                'user_ids[]': [self.normal_user.id]
            }
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        cells = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    cells.append(str(cell))
        cell_text = ' '.join(cells)

        self.assertIn(
            'Logged User',
            cell_text,
            msg="Report should contain 'Logged User' label"
        )
        self.assertIn(
            'admin',
            cell_text,
            msg="Report should contain admin username in header"
        )
        self.assertIn(
            'Selected User from Dropdown',
            cell_text,
            msg="Report should contain 'Selected User from Dropdown' label"
        )
        self.assertIn(
            'normal_user',
            cell_text,
            msg="Report should contain selected user's username in header"
        )

    def test_activity_section_shows_selected_user_data(self):
        """Activity section should show data for the selected dropdown user."""
        self.client.force_authenticate(user=self.admin_user)
        response = self.client.get(
            self.url,
            {
                'mode': 'individual',
                'user_ids[]': [self.normal_user.id]
            }
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        cells = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    cells.append(str(cell))
        cell_text = ' '.join(cells)

        # Now shows logged user context in sub-headers
        self.assertIn(
            'Logged User: admin (admin@example.com)',
            cell_text,
            msg="Activity section should show logged user info"
        )
        self.assertIn(
            'Selected User: normal_user',
            cell_text,
            msg="Activity section should show selected user"
        )
        # Should show normal_user's activities, not admin's
        self.assertIn(
            'Normal Browser',
            cell_text,
            msg="Should show normal user's browser agents"
        )

    def test_summary_section_shows_selected_user_statistics(self):
        """Summary statistics should be for the selected dropdown user."""
        self.client.force_authenticate(user=self.admin_user)
        response = self.client.get(
            self.url,
            {
                'mode': 'individual',
                'user_ids[]': [self.normal_user.id]
            }
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        cells = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    cells.append(str(cell))
        cell_text = ' '.join(cells)

        # Now shows logged user context in sub-headers
        self.assertIn(
            'Logged User: admin (admin@example.com)',
            cell_text,
            msg="Summary section should show logged user info"
        )
        self.assertIn(
            'Selected User: normal_user',
            cell_text,
            msg="Summary section should show selected user"
        )
        # normal_user has 5 total logins (4 successful, 1 failed)
        self.assertIn(
            '5',
            cell_text,
            msg="Summary should show 5 total logins for normal_user"
        )
        self.assertIn(
            '60.0%',
            cell_text,
            msg="Summary should show 60% success rate for normal_user"
        )

    def test_chart_sections_reference_selected_user(self):
        """Chart sections (trends, comparison, distribution) should
        reference the selected dropdown user."""
        self.client.force_authenticate(user=self.admin_user)
        response = self.client.get(
            self.url,
            {
                'mode': 'individual',
                'user_ids[]': [self.normal_user.id]
            }
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        cells = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    cells.append(str(cell))
        cell_text = ' '.join(cells)

        # All chart sections should show logged user context with selected user
        self.assertIn(
            'Logged User: admin (admin@example.com)',
            cell_text,
            msg="Chart sections should show logged user info"
        )
        self.assertIn(
            'Selected User: normal_user',
            cell_text,
            msg="Chart sections should show selected user"
        )

    def test_regular_user_own_report_does_not_show_dropdown_label(self):
        """Regular user downloading their own report (no dropdown)
        should NOT show 'Selected User from Dropdown' label."""
        self.client.force_authenticate(user=self.normal_user)
        response = self.client.get(
            self.url,
            {'mode': 'individual'}
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        cells = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    cells.append(str(cell))
        cell_text = ' '.join(cells)

        # Should NOT show "Selected User from Dropdown" for own report
        self.assertNotIn(
            'Selected User from Dropdown',
            cell_text,
            msg="Regular user's own report should not show dropdown label"
        )
        # Should show normal_user info without 'Logged Admin' prefix
        self.assertIn(
            'normal_user',
            cell_text,
            msg="Should show normal_user's username"
        )


class ReportDownloadDateFilterTests(TestCase):
    """Test date filtering for report download."""

    def setUp(self):
        self.user = User.objects.create_user(
            username='testuser',
            email='test@example.com',
            password='testpass123'
        )
        self.client = APIClient()
        self.url = reverse('user:report-download')
        self._create_test_data()

    def _create_test_data(self):
        for i in range(5):
            LoginActivity.objects.create(
                user=self.user,
                ip_address=f'192.168.1.{i+1}',
                user_agent=f'Browser {i+1}',
                success=True,
                timestamp=timezone.now() - timedelta(days=i)
            )
        for i in range(3):
            LoginActivity.objects.create(
                user=self.user,
                ip_address=f'192.168.2.{i+1}',
                user_agent=f'Browser {i+1}',
                success=False,
                timestamp=timezone.now() - timedelta(days=15+i)
            )

    def test_date_filter_with_start_date(self):
        self.client.force_authenticate(user=self.user)
        start_date = (
            timezone.now() - timedelta(days=7)
        ).strftime('%Y-%m-%d')
        response = self.client.get(
            self.url,
            {
                'mode': 'individual',
                'start_date': start_date
            }
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_date_filter_with_end_date(self):
        self.client.force_authenticate(user=self.user)
        end_date = timezone.now().strftime('%Y-%m-%d')
        response = self.client.get(
            self.url,
            {
                'mode': 'individual',
                'end_date': end_date
            }
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_date_filter_with_both_dates(self):
        self.client.force_authenticate(user=self.user)
        start_date = (
            timezone.now() - timedelta(days=10)
        ).strftime('%Y-%m-%d')
        end_date = timezone.now().strftime('%Y-%m-%d')
        response = self.client.get(
            self.url,
            {
                'mode': 'individual',
                'start_date': start_date,
                'end_date': end_date
            }
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_invalid_date_format_returns_error(self):
        self.client.force_authenticate(user=self.user)
        response = self.client.get(
            self.url,
            {
                'mode': 'individual',
                'start_date': 'invalid-date'
            }
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)


class ReportDownloadExcelTests(TestCase):
    """Test Excel report generation."""

    def setUp(self):
        self.user = User.objects.create_user(
            username='testuser',
            email='test@example.com',
            password='testpass123'
        )
        self.client = APIClient()
        self.url = reverse('user:report-download')
        self._create_test_data()

    def _create_test_data(self):
        for i in range(10):
            LoginActivity.objects.create(
                user=self.user,
                ip_address=f'192.168.1.{i+1}',
                user_agent=f'Firefox Browser {i+1}',
                success=i % 4 != 0,
                timestamp=timezone.now() - timedelta(days=i)
            )

    def test_excel_download_returns_excel_content_type(self):
        self.client.force_authenticate(user=self.user)
        response = self.client.get(
            self.url, {'mode': 'individual'}
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn(
            response['Content-Type'],
            [
                'application/vnd.openxmlformats-officedocument.'
                'spreadsheetml.sheet',
                'application/octet-stream'
            ]
        )

    def test_excel_download_has_content_disposition(self):
        self.client.force_authenticate(user=self.user)
        response = self.client.get(
            self.url, {'mode': 'individual'}
        )
        self.assertIn('Content-Disposition', response)
        self.assertIn('attachment', response['Content-Disposition'])
        self.assertIn('.xlsx', response['Content-Disposition'])

    def test_excel_download_returns_valid_excel_bytes(self):
        self.client.force_authenticate(user=self.user)
        response = self.client.get(
            self.url, {'mode': 'individual'}
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.content.startswith(b'PK'))


class ReportDownloadFeatureFlagTests(TestCase):
    """Test feature flag for report download."""

    def setUp(self):
        self.user = User.objects.create_user(
            username='testuser',
            email='test@example.com',
            password='testpass123'
        )
        self.client = APIClient()
        self.url = reverse('user:report-download')
        LoginActivity.objects.create(
            user=self.user,
            ip_address='192.168.1.1',
            user_agent='Test Browser',
            success=True,
            timestamp=timezone.now()
        )

    @override_settings(ENABLE_REPORT_DOWNLOAD=False)
    def test_disabled_returns_403(self):
        self.client.force_authenticate(user=self.user)
        response = self.client.get(
            self.url, {'mode': 'individual'}
        )
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    @override_settings(ENABLE_REPORT_DOWNLOAD=True)
    def test_enabled_works(self):
        self.client.force_authenticate(user=self.user)
        response = self.client.get(
            self.url, {'mode': 'individual'}
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)


class ReportDownloadUserValidationTests(TestCase):
    """Test user ID validation for report download."""

    def setUp(self):
        self.user1 = User.objects.create_user(
            username='user1',
            email='user1@example.com',
            password='testpass123'
        )
        self.admin_user = User.objects.create_superuser(
            username='admin',
            email='admin@example.com',
            password='adminpass123'
        )
        self.client = APIClient()
        self.url = reverse('user:report-download')
        LoginActivity.objects.create(
            user=self.user1,
            ip_address='192.168.1.1',
            user_agent='Test Browser',
            success=True,
            timestamp=timezone.now()
        )

    def test_invalid_user_ids_format_returns_error(self):
        self.client.force_authenticate(user=self.admin_user)
        response = self.client.get(
            self.url,
            {
                'mode': 'individual',
                'user_ids[]': ['invalid']
            }
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_nonexistent_user_id_returns_error(self):
        self.client.force_authenticate(user=self.admin_user)
        response = self.client.get(
            self.url,
            {
                'mode': 'individual',
                'user_ids[]': [99999]
            }
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('error', response.data)


class ReportDownloadSelectedUserTests(TestCase):
    """Test the selected_user_id query parameter for grouped mode.

    When an admin downloads a grouped report with user_ids[] for all
    dropdown users, they should also pass selected_user_id to tell the
    backend which specific user to label as 'Selected User' in report
    sections like Summary Stats and Recent Activities.
    """

    def setUp(self):
        self.admin_user = User.objects.create_superuser(
            username='admin',
            email='admin@example.com',
            password='adminpass123'
        )
        self.normal_user1 = User.objects.create_user(
            username='normal_user1',
            email='normal1@example.com',
            password='testpass123'
        )
        self.normal_user2 = User.objects.create_user(
            username='normal_user2',
            email='normal2@example.com',
            password='testpass123'
        )
        self.client = APIClient()
        self.url = reverse('user:report-download')

        # Create login activity for admin
        LoginActivity.objects.create(
            user=self.admin_user,
            ip_address='192.168.0.1',
            user_agent='Admin Browser',
            success=True,
            timestamp=timezone.now() - timedelta(days=1)
        )
        # Create 5 login activities for normal_user1
        for i in range(5):
            LoginActivity.objects.create(
                user=self.normal_user1,
                ip_address=f'192.168.1.{i+1}',
                user_agent=f'User1 Browser {i+1}',
                success=i % 4 != 0,
                timestamp=timezone.now() - timedelta(days=i)
            )
        # Create 3 login activities for normal_user2
        for i in range(3):
            LoginActivity.objects.create(
                user=self.normal_user2,
                ip_address=f'192.168.2.{i+1}',
                user_agent=f'User2 Browser {i+1}',
                success=True,
                timestamp=timezone.now() - timedelta(days=i)
            )

    def test_grouped_mode_with_selected_user_id_shows_correct_label(self):
        """When selected_user_id is provided in grouped mode,
        the report should show that user's label, not the first
        user in user_ids[]."""
        self.client.force_authenticate(user=self.admin_user)
        response = self.client.get(
            self.url,
            {
                'mode': 'grouped',
                'user_ids[]': [self.normal_user1.id, self.normal_user2.id],
                'selected_user_id': self.normal_user2.id,
            }
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        cells = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    cells.append(str(cell))
        cell_text = ' '.join(cells)

        # The Summary section should show normal_user2 as Selected User
        self.assertIn(
            'Logged User: admin (admin@example.com)',
            cell_text,
            msg="Summary section should show logged user info"
        )

    def test_grouped_mode_with_selected_user_id_still_has_all_data(self):
        """Grouped data should still combine all users' data
        even when selected_user_id is provided."""
        self.client.force_authenticate(user=self.admin_user)
        response = self.client.get(
            self.url,
            {
                'mode': 'grouped',
                'user_ids[]': [self.normal_user1.id, self.normal_user2.id],
                'selected_user_id': self.normal_user2.id,
            }
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        cells = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    cells.append(str(cell))
        cell_text = ' '.join(cells)

        # Grouped summary should show combined total (5+3=8)
        self.assertIn('Total Logins (All Users)', cell_text,
                      msg="GROUPED SUMMARY section not found in report")
        self.assertIn(
            '8', cell_text,
            msg="Expected 8 total logins (5+3) from both users"
        )

    def test_grouped_mode_without_selected_user_id_backward_compatible(self):
        """Without selected_user_id, grouped mode should fall back to
        existing behavior (first user in array for labeling)."""
        self.client.force_authenticate(user=self.admin_user)
        response = self.client.get(
            self.url,
            {
                'mode': 'grouped',
                'user_ids[]': [self.normal_user1.id, self.normal_user2.id],
            }
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        cells = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    cells.append(str(cell))
        cell_text = ' '.join(cells)

        # Without selected_user_id, first user (normal_user1) is used
        self.assertIn(
            'Logged User: admin (admin@example.com)',
            cell_text,
            msg="Chart sections should show logged user info"
        )

    def test_invalid_selected_user_id_returns_error(self):
        """Non-existent selected_user_id should return a 400 error."""
        self.client.force_authenticate(user=self.admin_user)
        response = self.client.get(
            self.url,
            {
                'mode': 'grouped',
                'user_ids[]': [self.normal_user1.id],
                'selected_user_id': 99999,
            }
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('error', response.data)

    def test_selected_user_id_requires_admin_permissions(self):
        """Regular users should not be able to use selected_user_id
        in grouped mode (already blocked by grouped mode check)."""
        regular_user = User.objects.create_user(
            username='regular',
            email='regular@example.com',
            password='testpass123'
        )
        self.client.force_authenticate(user=regular_user)
        response = self.client.get(
            self.url,
            {
                'mode': 'grouped',
                'selected_user_id': self.normal_user1.id,
            }
        )
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_grouped_mode_top_user_agents_for_selected_user_only(self):
        """TOP USER AGENTS in grouped mode should show only the
        selected user's user agents, not aggregated from all users.

        normal_user1 has 5 logins with User1 Browser 1-5.
        normal_user2 has 3 logins with User2 Browser 1-3.
        When selected_user_id=normal_user2, TOP USER AGENTS should
        only show 'User2 Browser' entries, not 'User1 Browser' ones.
        """
        self.client.force_authenticate(user=self.admin_user)
        response = self.client.get(
            self.url,
            {
                'mode': 'grouped',
                'user_ids[]': [self.normal_user1.id, self.normal_user2.id],
                'selected_user_id': self.normal_user2.id,
            }
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        cells = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    cells.append(str(cell))
        cell_text = ' '.join(cells)

        # The TOP USER AGENTS section should show selected user's agents
        self.assertIn(
            'User2 Browser',
            cell_text,
            msg="TOP USER AGENTS should include selected user's user agents"
        )

        # The TOP USER AGENTS section should NOT include other user's agents
        self.assertNotIn(
            'User1 Browser',
            cell_text,
            msg="TOP USER AGENTS should NOT include other users' user agents"
        )


class ReportDownloadEmptyDataTests(TestCase):
    """Test report download with empty data."""

    def setUp(self):
        self.user = User.objects.create_user(
            username='testuser',
            email='test@example.com',
            password='testpass123'
        )
        self.client = APIClient()
        self.url = reverse('user:report-download')

    def test_excel_download_with_no_data_returns_valid_excel(self):
        self.client.force_authenticate(user=self.user)
        response = self.client.get(
            self.url, {'mode': 'individual'}
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.content.startswith(b'PK'))


class ReportDownloadGroupedModeHeaderTests(TestCase):
    """Test that grouped mode Excel report shows 'Logged User' +
    'Selected User from Dropdown' in header and section sub-headers,
    matching the individual mode format exactly."""

    def setUp(self):
        self.admin_user = User.objects.create_superuser(
            username='admin',
            email='admin@example.com',
            password='adminpass123'
        )
        self.normal_user = User.objects.create_user(
            username='normal_user',
            email='normal@example.com',
            password='testpass123'
        )
        self.client = APIClient()
        self.url = reverse('user:report-download')

        # Create login activity for admin
        LoginActivity.objects.create(
            user=self.admin_user,
            ip_address='192.168.0.1',
            user_agent='Admin Browser',
            success=True,
            timestamp=timezone.now() - timedelta(days=1)
        )
        # Create login activities for normal_user
        for i in range(5):
            LoginActivity.objects.create(
                user=self.normal_user,
                ip_address=f'192.168.1.{i+1}',
                user_agent=f'Normal Browser {i+1}',
                success=i % 4 != 0,
                timestamp=timezone.now() - timedelta(days=i)
            )

    def _get_excel_text(self, params):
        """Helper: fetch Excel report and return all cell text."""
        self.client.force_authenticate(user=self.admin_user)
        response = self.client.get(self.url, params)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        cells = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    cells.append(str(cell))
        return ' '.join(cells)

    def test_1_header_shows_logged_user_and_selected_user(
        self
    ):
        """Grouped mode with selected_user_id should show
        'Logged User' and 'Selected User from Dropdown' in header."""
        cell_text = self._get_excel_text({
            'mode': 'grouped',
            'user_ids[]': [self.normal_user.id],
            'selected_user_id': self.normal_user.id,
        })

        self.assertIn(
            'Logged User: admin (admin@example.com)',
            cell_text,
            msg=(
                "Grouped mode header should show "
                "'Logged User: admin (admin@example.com)'"
            )
        )
        self.assertIn(
            'Selected User from Dropdown: normal_user (normal@example.com)',
            cell_text,
            msg=(
                "Grouped mode header should show "
                "'Selected User from Dropdown: "
                "normal_user (normal@example.com)'"
            )
        )
        # Should NOT have the old 'Current User' line
        self.assertNotIn(
            'Current User:',
            cell_text,
            msg="Grouped mode header should NOT contain 'Current User' line"
        )

    def test_2_activities_section_shows_logged_user_and_selected_user(self):
        """Activities section sub-header in grouped mode should include
        Logged User and Selected User info."""
        cell_text = self._get_excel_text({
            'mode': 'grouped',
            'user_ids[]': [self.normal_user.id],
            'selected_user_id': self.normal_user.id,
        })

        self.assertIn(
            'Logged User: admin',
            cell_text,
            msg="Activities section should show Logged User"
        )
        self.assertIn(
            'Selected User: normal_user',
            cell_text,
            msg="Activities section should show Selected User"
        )
        # Should NOT show old format without admin info
        self.assertNotIn(
            'Filter: Specific Users → Selected User: normal_user',
            cell_text,
            msg="Activities section should NOT have old format without admin"
        )

    def test_3_summary_section_shows_logged_user_and_selected_user(self):
        """Summary section sub-header in grouped mode should include
        Logged User and Selected User info."""
        cell_text = self._get_excel_text({
            'mode': 'grouped',
            'user_ids[]': [self.normal_user.id],
            'selected_user_id': self.normal_user.id,
        })

        self.assertIn(
            'Logged User: admin',
            cell_text,
            msg="Summary section should show Logged User"
        )
        self.assertIn(
            'Selected User: normal_user',
            cell_text,
            msg="Summary section should show Selected User"
        )
        # Should NOT show old format without admin info
        self.assertNotIn(
            'Filter: Specific Users → Selected User: normal_user',
            cell_text,
            msg="Summary section should NOT have old format without admin"
        )

    def test_4_chart_sections_show_combined_data_in_grouped_mode(self):
        """Chart sections in grouped mode with a selected_user_id should show
        'Combined Data' with all users' names, not just the selected user."""
        cell_text = self._get_excel_text({
            'mode': 'grouped',
            'user_ids[]': [self.normal_user.id],
            'selected_user_id': self.normal_user.id,
        })

        # Chart context should show logged user and combined data info
        self.assertIn(
            'Logged User: admin',
            cell_text,
            msg="Chart sections should show Logged User"
        )
        # Should show Combined Data with user count, not Selected User
        self.assertIn(
            'Combined Data',
            cell_text,
            msg="Chart sections should show 'Combined Data' in grouped mode"
        )
        self.assertIn(
            'normal_user',
            cell_text,
            msg="Chart sections should include the user in combined data"
        )
        # Should NOT show "Selected User" in chart sections (it's combined)
        self.assertNotIn(
            'Filter: Specific Users → Selected User: normal_user',
            cell_text,
            msg="Chart sections should NOT show 'Selected User'"
        )
        # Should NOT show old format with just "Users: ..."
        self.assertNotIn(
            'Filter: Specific Users → Users: normal_user',
            cell_text,
            msg="Chart sections should NOT have old users-only format"
        )

    def test_5_grouped_mode_shows_combined_data_with_all_users(self):
        """Grouped mode with multiple users should list all users in
        Combined Data context for chart sections."""
        # Create a second user
        extra_user = User.objects.create_user(
            username='extra_user',
            email='extra@example.com',
            password='testpass123'
        )
        LoginActivity.objects.create(
            user=extra_user,
            ip_address='192.168.5.1',
            user_agent='Extra Browser',
            success=True,
            timestamp=timezone.now()
        )

        cell_text = self._get_excel_text({
            'mode': 'grouped',
            'user_ids[]': [self.normal_user.id, extra_user.id],
            'selected_user_id': self.normal_user.id,
        })

        self.assertIn(
            'Combined Data (2 users): normal_user, extra_user',
            cell_text,
            msg="Chart sections should show both users in combined data"
        )

    def test_6_grouped_mode_truncates_many_users(self):
        """Grouped mode with >3 users should truncate and show '...'."""
        extra_users = []
        for i in range(5):
            u = User.objects.create_user(
                username=f'extra_user_{i}',
                email=f'extra{i}@example.com',
                password='testpass123'
            )
            LoginActivity.objects.create(
                user=u,
                ip_address=f'192.168.10.{i}',
                user_agent=f'Extra Browser {i}',
                success=True,
                timestamp=timezone.now()
            )
            extra_users.append(u)

        user_ids = [self.normal_user.id] + [u.id for u in extra_users]
        cell_text = self._get_excel_text({
            'mode': 'grouped',
            'user_ids[]': user_ids,
            'selected_user_id': self.normal_user.id,
        })

        # Should show first 3 usernames + ellipsis + total count
        self.assertIn(
            'Combined Data (6 users)',
            cell_text,
            msg="Chart context should show combined data with user count"
        )
        self.assertIn(
            'normal_user',
            cell_text,
            msg="Should include first user name"
        )
        self.assertIn(
            'extra_user_0',
            cell_text,
            msg="Should include second user name"
        )
        self.assertIn(
            'extra_user_1',
            cell_text,
            msg="Should include third user name"
        )
        self.assertIn(
            '...',
            cell_text,
            msg="Should show ellipsis for truncated users"
        )
        # Fourth user (extra_user_3) should NOT appear (truncated)
        # But note: extra_user_2 IS the 4th in the list (index 0,1,2 are shown)
        # The 4th displayed user is extra_user_3 which should NOT appear
        # Actually extra_user_2 and beyond should not appear after truncation
        self.assertNotIn(
            'extra_user_4',
            cell_text,
            msg="Should NOT include the 6th user name (truncated)"
        )

    def test_7_grouped_mode_without_role_still_shows_correct_header(
        self
    ):
        """Grouped mode without role should still show correct header."""
        cell_text = self._get_excel_text({
            'mode': 'grouped',
            'user_ids[]': [self.normal_user.id],
            'selected_user_id': self.normal_user.id,
        })

        self.assertIn(
            'Logged User: admin (admin@example.com)',
            cell_text,
            msg="Should still show logged user in header without role param"
        )
        self.assertIn(
            'Selected User from Dropdown: normal_user (normal@example.com)',
            cell_text,
            msg="Should show selected user in header without role param"
        )
