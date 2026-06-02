"""Excel report generator for login tracking dashboard."""
import io

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


class ExcelReportGenerator:
    """Generates Excel reports for login tracking data (single sheet)."""

    def __init__(self, report_data):
        self.data = report_data
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Login Activity Report"
        self._setup_styles()

    def _setup_styles(self):
        self.title_font = Font(
            name='Arial', size=18, bold=True, color='1976D2'
        )
        self.header_font = Font(
            name='Arial', size=14, bold=True, color='424242'
        )
        self.subheader_font = Font(
            name='Arial', size=12, bold=True, color='616161'
        )
        self.label_font = Font(name='Arial', size=10, bold=True)
        self.value_font = Font(name='Arial', size=10)
        self.metadata_font = Font(name='Arial', size=9, color='757575')
        self.username_font = Font(
            name='Arial', size=10, bold=True, color='37474F'
        )
        self.header_fill = PatternFill(
            start_color='1976D2', end_color='1976D2', fill_type='solid'
        )
        self.header_font_white = Font(
            name='Arial', size=10, bold=True, color='FFFFFF'
        )
        self.alt_row_fill = PatternFill(
            start_color='F5F5F5', end_color='F5F5F5', fill_type='solid'
        )
        self.success_fill = PatternFill(
            start_color='E8F5E9', end_color='E8F5E9', fill_type='solid'
        )
        self.failure_fill = PatternFill(
            start_color='FFEBEE', end_color='FFEBEE', fill_type='solid'
        )
        self.center_align = Alignment(
            horizontal='center', vertical='center'
        )
        self.left_align = Alignment(horizontal='left', vertical='center')
        self.wrap_align = Alignment(
            horizontal='left', vertical='center', wrap_text=True
        )
        self.thin_border = Border(
            left=Side(style='thin', color='E0E0E0'),
            right=Side(style='thin', color='E0E0E0'),
            top=Side(style='thin', color='E0E0E0'),
            bottom=Side(style='thin', color='E0E0E0')
        )

    def generate(self):
        """Order: Header -> Activities -> User Agents -> Summary -> Charts"""
        self._mode_label = (
            'Grouped' if self.data['mode'] == 'grouped' else 'Individual'
        )
        row = 1
        row = self._build_header(row)
        row += 2
        row = self._build_activities_section(row)
        row += 2
        row = self._build_user_agents_section(row)
        row += 2
        row = self._build_summary_section(row)
        row += 2
        row = self._build_trends_section(row)
        row += 2
        row = self._build_comparison_section(row)
        row += 2
        row = self._build_distribution_section(row)
        row += 2
        row = self._build_grouped_summary_section(row)
        self._adjust_column_widths()
        buf = io.BytesIO()
        self.wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    def _strftime_safe(self, dt, fmt='%Y-%m-%d'):
        if hasattr(dt, 'strftime'):
            return dt.strftime(fmt)
        return str(dt)[:10] if dt else 'N/A'

    def _format_user_list(self, user_details=None, max_display=3):
        """Format a list of usernames with truncation for large groups.

        Args:
            user_details: List of user detail dicts with 'username' keys.
                          If None, uses self.data['user_details'].
            max_display: Maximum number of usernames to show before
                         truncating with '...'.

        Returns:
            tuple: (formatted_names_string, total_count)
        """
        if user_details is None:
            user_details = self.data.get('user_details', [])
        count = len(user_details)
        names = [u['username'] for u in user_details]

        if count <= max_display:
            return ', '.join(names), count
        else:
            first_few = ', '.join(names[:max_display])
            return f"{first_few}, ... ({count} users)", count

    def _get_selected_username(self):
        """Get the selected user's display name."""
        if hasattr(self.data.get('user'), 'username'):
            return self.data['user'].username
        details = self.data.get('user_details', [])
        return details[0]['username'] if details else 'N/A'

    def _has_dropdown_selection(self):
        """Check if this report is for a dropdown-selected user by an admin."""
        return self.data.get('has_dropdown_selection', False)

    def _write_table_header(self, row, headers):
        for j, h in enumerate(headers):
            c = self.ws.cell(row=row, column=j+1, value=h)
            c.font = self.header_font_white
            c.fill = self.header_fill
            c.border = self.thin_border
        return row + 1

    def _write_section_title(self, row, title):
        self.ws.cell(row=row, column=1, value=title).font = self.header_font
        return row + 1

    def _get_filter_label(self):
        """Get human-readable filter label from filter_info."""
        filter_info = self.data.get('filter_info', {})
        ft = filter_info.get('type')
        role = filter_info.get('role')

        # Map filter type to display label
        label_map = {
            'all': 'All Users',
            'admin_only': 'Admin Only',
            'regular_users': 'Regular Users',
            'me': 'My Data',
        }

        if ft in label_map:
            label = label_map[ft]
        elif ft == 'user_ids':
            label = 'Specific Users'
        else:
            label = 'No Filter'

        # Append role if present
        if role:
            role_label = 'Admin' if role == 'admin' else 'Regular'
            label += f' (Role: {role_label})'

        return label

    def _build_subheader_with_admin_context(
        self, filter_label, selected_user_label
    ):
        """Build section sub-header with admin info when applicable.

        For grouped/individual mode with a dropdown selection, includes
        both the Logged User and Selected User info in the label.
        """
        if self._has_dropdown_selection():
            requesting_user = self.data.get('requesting_user')
            if requesting_user and hasattr(requesting_user, 'username'):
                logged_user_info = (
                    f"Logged User: {requesting_user.username} "
                    f"({requesting_user.email})"
                )
                return (
                    f"Filter: {filter_label} → "
                    f"{logged_user_info} → "
                    f"Selected User: {selected_user_label}"
                )

        # Fall back to standard format
        return f"Filter: {filter_label} → Selected User: {selected_user_label}"

    def _write_chart_context(self, row):
        """Write filter info + users under chart title."""
        filter_label = self._get_filter_label()

        if self._has_dropdown_selection():
            requesting_user = self.data.get('requesting_user')
            if requesting_user and hasattr(requesting_user, 'username'):
                # For grouped mode with dropdown selection, show combined data
                user_names, user_count = self._format_user_list()
                logged_user_info = (
                    f"Logged User: {requesting_user.username} "
                    f"({requesting_user.email})"
                )
                ctx = (
                    f"Filter: {filter_label} → "
                    f"{logged_user_info} → "
                    f"Combined Data ({user_count} users): {user_names}"
                )
            else:
                # Fallback - should not normally reach here
                selected_user = self._get_selected_username()
                ctx = self._build_subheader_with_admin_context(
                    filter_label, selected_user
                )
        elif self._mode_label == 'Grouped':
            user_names, user_count = self._format_user_list()
            ctx = (
                f"Filter: {filter_label} → "
                f"Combined Data ({user_count} users): {user_names}"
            )
        else:
            selected_user = self._get_selected_username()
            ctx = f"Filter: {filter_label} → Selected User: {selected_user}"

        self.ws.cell(row=row, column=1, value=ctx).font = Font(
            name='Arial', size=9, italic=True, color='78909C'
        )
        return row + 1

    def _build_header(self, start_row):
        row = start_row
        self.ws.cell(
            row=row, column=1,
            value="LOGIN ACTIVITY SUMMARY REPORT"
        ).font = self.title_font
        row += 1

        has_dropdown = self._has_dropdown_selection()
        requesting_user = self.data.get('requesting_user')

        if has_dropdown and requesting_user:
            # Admin selected a dropdown user - show both
            logged_user_info = (
                f"Logged User: {requesting_user.username} "
                f"({requesting_user.email})"
            )
            self.ws.cell(
                row=row, column=1, value=logged_user_info
            ).font = self.username_font
            row += 1

            selected_user = self.data.get('user')
            if hasattr(selected_user, 'username'):
                user_info = (
                    "Selected User from Dropdown: "
                    f"{selected_user.username} "
                    f"({selected_user.email})"
                )
                self.ws.cell(
                    row=row, column=1, value=user_info
                ).font = Font(
                    name='Arial', size=10, bold=True, color='1565C0'
                )
                row += 1
        else:
            # Standard display - current user
            if self.data.get('user'):
                ui = (
                    f"Current User: {self.data['user'].username} "
                    f"({self.data['user'].email})"
                )
            else:
                ud = self.data.get('user_details', [])
                first = ud[0] if ud else {}
                ui = f"Current User: {first.get('username', 'N/A')}"
                if len(ud) > 1:
                    names = ', '.join([u['username'] for u in ud])
                    ui += f" | All: {names}"
            self.ws.cell(
                row=row, column=1, value=ui
            ).font = self.username_font
            row += 1

        s = self._strftime_safe(self.data['start_date'])
        e = self._strftime_safe(self.data['end_date'])
        self.ws.cell(
            row=row, column=1,
            value=f"Period: {s} to {e}"
        ).font = self.metadata_font
        row += 1

        gen = self._strftime_safe(
            self.data['generated_at'], '%Y-%m-%d %H:%M:%S'
        )
        self.ws.cell(
            row=row, column=1,
            value=f"Generated: {gen}"
        ).font = self.metadata_font
        row += 1
        return row

    def _build_activities_section(self, start_row):
        row = start_row
        row = self._write_section_title(row, "RECENT LOGIN ACTIVITIES")

        uname = self._get_selected_username()
        filter_label = self._get_filter_label()
        label = self._build_subheader_with_admin_context(filter_label, uname)
        self.ws.cell(
            row=row, column=1, value=label
        ).font = Font(
            name='Arial', size=9, italic=True, color='78909C'
        )
        row += 1

        activities = self.data.get('login_activities', [])
        if not activities:
            self.ws.cell(
                row=row, column=1,
                value="No login activities found for the period."
            ).font = self.value_font
            return row + 1

        row = self._write_table_header(
            row, ['Timestamp', 'Username', 'IP Address', 'User Agent',
                  'Status']
        )
        for i, a in enumerate(activities):
            status = 'Success' if a['success'] else 'Failed'
            dr = [a['timestamp'], a['username'], a['ip_address'],
                  a['user_agent'], status]
            for j, val in enumerate(dr):
                c = self.ws.cell(row=row, column=j+1, value=val)
                c.font = self.value_font
                c.border = self.thin_border
                if j == 3:
                    c.alignment = self.wrap_align
                if j == 4:
                    c.fill = (self.success_fill if a['success']
                              else self.failure_fill)
                elif i % 2 == 1:
                    c.fill = self.alt_row_fill
            row += 1
        return row

    def _build_user_agents_section(self, start_row):
        """Top User Agents shown right after activities."""
        row = start_row
        row = self._write_section_title(row, "TOP USER AGENTS")

        dd = self.data.get('login_distribution', {})
        ua = dd.get('user_agents', {}) if dd else {}

        if not ua or not ua.get('labels') or not ua.get('datasets'):
            self.ws.cell(
                row=row, column=1,
                value="No user agent data available."
            ).font = self.value_font
            return row + 1

        labels = ua['labels']
        data_values = ua['datasets'][0]['data'] if ua['datasets'] else []

        row = self._write_table_header(row, ['User Agent', 'Count'])
        for i, label in enumerate(labels):
            count = data_values[i] if i < len(data_values) else 0
            for j, val in enumerate([label, count]):
                c = self.ws.cell(row=row, column=j+1, value=val)
                c.font = self.value_font
                c.border = self.thin_border
                if j == 0:
                    c.alignment = self.wrap_align
                if i % 2 == 1:
                    c.fill = self.alt_row_fill
            row += 1
        return row

    def _build_summary_section(self, start_row):
        row = start_row
        row = self._write_section_title(row, "SUMMARY STATISTICS")

        uname = self._get_selected_username()
        filter_label = self._get_filter_label()
        label = self._build_subheader_with_admin_context(filter_label, uname)
        self.ws.cell(
            row=row, column=1, value=label
        ).font = Font(
            name='Arial', size=9, italic=True, color='78909C'
        )
        row += 1

        s = self.data['summary']
        total = s.get('total_logins', 0)
        ok_ = s.get('total_successful_logins', 0)
        rate = f"{(ok_ / total * 100):.1f}%" if total > 0 else "N/A"

        data = [
            ['Metric', 'Value'],
            ['Total Logins', str(s.get('total_logins', 0))],
            ['Successful Logins', str(ok_)],
            ['Failed Logins', str(s.get('total_failed_logins', 0))],
            ['Success Rate', rate],
            ['Last Login', str(s.get('last_login', 'N/A'))],
        ]
        for i, dr in enumerate(data):
            for j, val in enumerate(dr):
                c = self.ws.cell(row=row, column=j+1, value=val)
                c.border = self.thin_border
                if i == 0:
                    c.font = self.header_font_white
                    c.fill = self.header_fill
                else:
                    c.font = self.label_font if j == 0 else self.value_font
                    if i % 2 == 0:
                        c.fill = self.alt_row_fill
            row += 1
        return row

    def _build_trends_section(self, start_row):
        row = start_row
        row = self._write_section_title(
            row, f"LOGIN TRENDS DATA ({self._mode_label})"
        )
        row = self._write_chart_context(row)

        td = self.data.get('login_trends', {})
        if not td or not td.get('labels'):
            self.ws.cell(
                row=row, column=1,
                value="No login trend data available."
            ).font = self.value_font
            return row + 1

        labels = td['labels']
        datasets = td.get('datasets', [])
        s_data = []
        f_data = []
        for ds in datasets:
            if 'Successful' in ds['label']:
                s_data = ds['data']
            elif 'Failed' in ds['label']:
                f_data = ds['data']

        filtered = []
        for i, date in enumerate(labels):
            s = float(s_data[i]) if i < len(s_data) else 0
            f = float(f_data[i]) if i < len(f_data) else 0
            if s > 0 or f > 0:
                filtered.append((date, int(s), int(f)))

        if not filtered:
            self.ws.cell(
                row=row, column=1,
                value="No activity recorded in this period."
            ).font = self.value_font
            return row + 1

        row = self._write_table_header(
            row, ['Date', 'Successful Logins', 'Failed Logins']
        )
        for i, (date, s, f) in enumerate(filtered):
            for j, v in enumerate([date, s, f]):
                c = self.ws.cell(row=row, column=j+1, value=v)
                c.font = self.value_font
                c.border = self.thin_border
                if i % 2 == 1:
                    c.fill = self.alt_row_fill
            row += 1
        return row

    def _build_comparison_section(self, start_row):
        row = start_row
        row = self._write_section_title(
            row, f"LOGIN COMPARISON DATA ({self._mode_label})"
        )
        row = self._write_chart_context(row)

        cd = self.data.get('login_comparison', {})
        if not cd or not cd.get('labels'):
            self.ws.cell(
                row=row, column=1,
                value="No comparison data available."
            ).font = self.value_font
            return row + 1

        labels = cd['labels']
        datasets = cd.get('datasets', [])
        vals = datasets[0]['data'] if datasets else []

        filtered = []
        for i, label in enumerate(labels):
            v = vals[i] if i < len(vals) else 0
            if float(v) > 0:
                filtered.append((label, v))

        if not filtered:
            self.ws.cell(
                row=row, column=1,
                value="No comparison data available."
            ).font = self.value_font
            return row + 1

        row = self._write_table_header(row, ['Period', 'Login Count'])
        for i, (label, val) in enumerate(filtered):
            for j, v in enumerate([label, val]):
                c = self.ws.cell(row=row, column=j+1, value=v)
                c.font = self.value_font
                c.border = self.thin_border
                if i % 2 == 1:
                    c.fill = self.alt_row_fill
            row += 1
        return row

    def _build_distribution_section(self, start_row):
        row = start_row
        row = self._write_section_title(
            row, f"LOGIN DISTRIBUTION ({self._mode_label})"
        )
        row = self._write_chart_context(row)

        dd = self.data.get('login_distribution', {})
        if not dd:
            self.ws.cell(
                row=row, column=1,
                value="No distribution data available."
            ).font = self.value_font
            return row + 1

        sr = dd.get('success_ratio', {})
        self.ws.cell(
            row=row, column=1,
            value="Success/Failure Ratio"
        ).font = self.subheader_font
        row += 1

        if sr and sr.get('labels'):
            labels = sr['labels']
            datasets = sr.get('datasets', [])
            dv = datasets[0]['data'] if datasets else []
            total = sum(dv) if dv else 0
            row = self._write_table_header(
                row, ['Status', 'Count', 'Percentage']
            )
            for i, label in enumerate(labels):
                count = dv[i] if i < len(dv) else 0
                pct = (
                    f"{(count / total * 100):.1f}%"
                    if total > 0 else "0%"
                )
                for j, v in enumerate([label, count, pct]):
                    c = self.ws.cell(row=row, column=j+1, value=v)
                    c.font = self.value_font
                    c.border = self.thin_border
                    c.fill = (
                        self.success_fill
                        if label == 'Successful'
                        else self.failure_fill
                    )
                row += 1
        return row

    def _build_grouped_summary_section(self, start_row):
        """Build grouped summary like admin overview."""
        gs = self.data.get('grouped_summary')
        if not gs:
            return start_row

        row = start_row
        row = self._write_section_title(
            row, "GROUPED SUMMARY (Admin Overview)"
        )

        user_names, user_count = self._format_user_list()
        ctx = f"Combined statistics for {user_count} users: {user_names}"
        self.ws.cell(
            row=row, column=1, value=ctx
        ).font = Font(
            name='Arial', size=9, italic=True, color='78909C'
        )
        row += 1

        total = gs.get('total_logins', 0)
        ok_ = gs.get('total_successful_logins', 0)
        failed = gs.get('total_failed_logins', 0)
        rate = f"{(ok_ / total * 100):.1f}%" if total > 0 else "N/A"

        data = [
            ['Metric', 'Value'],
            ['Total Logins (All Users)', str(total)],
            ['Successful Logins (All Users)', str(ok_)],
            ['Failed Logins (All Users)', str(failed)],
            ['Success Rate (All Users)', rate],
        ]

        for i, dr in enumerate(data):
            for j, val in enumerate(dr):
                c = self.ws.cell(row=row, column=j+1, value=val)
                c.border = self.thin_border
                if i == 0:
                    c.font = self.header_font_white
                    c.fill = self.header_fill
                else:
                    c.font = self.label_font if j == 0 else self.value_font
                    if i % 2 == 0:
                        c.fill = self.alt_row_fill
            row += 1
        return row

    def _adjust_column_widths(self):
        widths = {'A': 28, 'B': 20, 'C': 20, 'D': 40, 'E': 15}
        for k, w in widths.items():
            self.ws.column_dimensions[k].width = w
